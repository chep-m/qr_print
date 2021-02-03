import openpyxl
import pyqrcode
import qrcode
import os 
import re


print("\nТекущая ссылка: http://url.ru/check/\n")

while True:
    try:
        strin = input(r'Введите путь к .xlsx файлу (C:/Users/.../example.xlsx):')
        strin=strin.replace('\\', "/")
        wb = openpyxl.load_workbook(strin)
        sheets = wb.sheetnames
        sheet = wb.active
        rows = sheet.max_row
        cols = sheet.max_column
        print()
        break
    except Exception:
        print("\n\tУпс! Похоже не удалось найти такой файл, попробуйте еще.\n")

n = 0

def qrc(site, pat):
    site = site[0: -5]
    filename = pat + '/' + site + '.png'
    site = 'http://url.ru/check/' + site
    img = qrcode.make(site)
    img.save(filename)

while True:
    try:
        data=input(r'Укажите папку, в которую сохранятся qr-коды (C:\Users\...\folder):')
        data = data.replace('\\','/')

        for i in range(1, rows + 1):
            for j in range(1, cols + 1):
                cell = sheet.cell(row=i, column=j)
                string =str(cell.value)
                if string == 'None' or string == ' ' or string == '  ' or string == '':
                    continue
                qrc(string,data)
                n += 1
        break

    except Exception:
        print("\n\tОй! Похоже такой папки нет, попробуйте еще.\n")
    
print('\nСоздано файлов: ', n)
input("\nНажмите Enter, чтобы выйти.")